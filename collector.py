from __future__ import annotations

import hashlib
import json
import math
import os
import random
import re
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, urlparse

import pandas as pd
import requests
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class MatchBindingInfo:
    input_url: str
    normalized_url: str
    event_key: str
    k: str
    token: str
    title: str
    start_time: str
    end_time: str
    wp_stop_time: str
    login_url: str
    situation_url: str
    problems_url: str
    bind_source: str
    entry_type: int
    mode_type: int
    score_type: int
    school_label: str
    industry_label: str


class ExportStatus:
    """导出状态枚举。"""

    SUCCESS = "success"      # 正常导出，有数据
    NO_DATA = "no_data"      # 接口返回"暂无数据"等业务级空结果
    FAILED = "failed"        # 调用过程中发生未捕获异常


class NoDataError(Exception):
    """业务级"无数据"异常：例如接口 code=116 message 含"暂无数据"。

    与真正错误的区别：这是平台告知"该榜单就是没有数据"，应跳过/写提示 Sheet，
    而不是以 RuntimeError 形式让调用方失败。
    """

    def __init__(self, board_name: str, message: str = "") -> None:
        self.board_name = board_name
        self.message = message or "暂无数据"
        super().__init__(f"{board_name}：{self.message}")


@dataclass
class LeaderboardExportResult:
    board_key: str
    board_name: str
    status: str  # ExportStatus 中的一个
    excel_path: str  # NO_DATA / FAILED 时可能为空字符串
    total_rows: int
    total_pages: int
    endpoint: str
    error_message: str = ""  # 仅 FAILED / NO_DATA 时填充
    debug_json_path: str = ""


class IChunQiuCollector:
    BASE_API = os.getenv("ICHUNQIU_BASE_API", "https://apiterminator.ichunqiu.com").rstrip("/")
    API_SECRET = os.getenv("ICHUNQIU_API_SECRET", "7637b08bdb0b29e08300a976b24ca672")

    # 浏览器 User-Agent 与来源信息，避免被 CDN/WAF 识别为脚本请求而拦截。
    USER_AGENT = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    )
    ORIGIN = "https://match.ichunqiu.com"

    BOARD_CONFIGS: Dict[str, Dict[str, str]] = {
        "solved": {"name": "解题总榜", "endpoint": "/match/rank/solved", "sheet": "解题总榜表"},
        "integral": {"name": "积分总榜", "endpoint": "/match/rank/team", "sheet": "积分总榜表"},
        "question": {"name": "题目榜单", "endpoint": "/match/rank/question", "sheet": "题目榜单表"},
        "single": {"name": "单人总榜", "endpoint": "/match/rank/person", "sheet": "单人总榜表"},
        "category": {"name": "题型榜单", "endpoint": "/match/rank/category", "sheet": "题型榜单表"},
        "team_info": {"name": "团队信息", "endpoint": "/match/team/users", "sheet": "团队成员"},
    }

    OFFICIAL_CONTENT: Dict[str, str] = {
        "pm": "排名",
        "teamName": "队伍名称",
        "xsmc": "选手名称",
        "school": "学校/单位名称",
        "sshy": "所属行业",
        "totalScore": "总分",
        "difference": "与前一名分差",
        "topicName": "题目名称",
        "topicType": "题目类型",
        "aqzszf": "理论知识总分",
        "aqzsjd": "理论知识总分",
        "jtsl": "解题数",
        "jcdtm": "解出的题目",
        "dwcy": "队伍成员",
        "grqx": "个人强项",
        "dwqx": "队伍强项",
        "province": "省份",
        "ssfq": "所属分区",
        "user_id": "用户ID",
        "user_name": "成员名称",
        "user_type": "成员角色",
        "team_id": "队伍ID",
        "team_name": "队伍名称",
        "team_rank": "排名",
        "is_reply": "是否参赛",
        "is_login": "是否签到",
        "is_upload_answer": "是否上传答案",
        "face_status": "人脸核验",
        "face_on_time": "人脸签到时间",
        "face_off_time": "人脸签退时间",
        "call_time": "呼叫时间",
        "micro_time": "在线时长(秒)",
        "create_time": "加入时间",
        "update_time": "更新时间",
    }

    @staticmethod
    def _role_label(user_type: Any) -> str:
        try:
            v = int(user_type)
        except (TypeError, ValueError):
            return str(user_type)
        if v == 1:
            return "队员"
        if v == 2:
            return "队长"
        return str(v)

    def __init__(self, log_callback: Optional[Callable[[str], None]] = None, lang: str = "zh-cn") -> None:
        self._log_callback = log_callback
        self._lang = lang
        self._session = requests.Session()
        self._request_logs: List[Dict[str, Any]] = []

    def _log(self, message: str) -> None:
        if self._log_callback:
            self._log_callback(message)

    @staticmethod
    def _now() -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _safe_name(text: str) -> str:
        cleaned = re.sub(r"[\\\\/:*?\"<>|]+", "_", text).strip()
        return cleaned or "match"

    @staticmethod
    def _safe_sheet_name(text: str, used_names: set[str]) -> str:
        cleaned = re.sub(r"[\\\\/:*?\[\]]+", "_", text).strip() or "Sheet"
        cleaned = cleaned[:31]
        name = cleaned
        idx = 2
        while name in used_names:
            suffix = f"_{idx}"
            name = f"{cleaned[: max(0, 31 - len(suffix))]}{suffix}"
            idx += 1
        used_names.add(name)
        return name

    @staticmethod
    def _parse_input_url(user_input: str) -> Tuple[str, str, str]:
        raw = user_input.strip()
        if not raw:
            raise ValueError("请输入比赛地址。")
        if not raw.startswith(("http://", "https://")):
            raw = f"https://match.ichunqiu.com/{raw.lstrip('/')}"

        parsed = urlparse(raw)
        if not parsed.netloc:
            raise ValueError("比赛地址格式不正确。")

        query = parse_qs(parsed.query)
        k_query = (query.get("k") or [""])[0]

        path = parsed.path.strip("/")
        generic_paths = {
            "login",
            "situation",
            "situation/problems",
            "situation/integral",
            "situation/question",
            "situation/single",
            "situation/category",
        }
        if path in generic_paths:
            event_key = ""
        else:
            event_key = path.split("/")[0] if path else ""

        return raw, event_key, k_query

    @staticmethod
    def _as_dict(value: Any) -> Dict[str, Any]:
        return value if isinstance(value, dict) else {}

    @staticmethod
    def _as_list(value: Any) -> List[Any]:
        return value if isinstance(value, list) else []

    @staticmethod
    def _to_int(value: Any, default: int = 0) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _calc_rs() -> str:
        seed = f"{time.time() * 1000 + random.random()}icq"
        return hashlib.md5(seed.encode("utf-8")).hexdigest()

    @classmethod
    def _calc_sign(cls, payload: Dict[str, Any]) -> str:
        parts = [f"{k}={payload[k]}" for k in payload]
        text = "&".join(sorted(parts)) + f"&{cls.API_SECRET}"
        return hashlib.sha1(text.encode("utf-8")).hexdigest()

    def _make_payload(self, data: Dict[str, Any], k: str = "", token: str = "") -> Dict[str, Any]:
        payload = dict(data)
        payload["k"] = payload.get("k", k)
        payload["stamp"] = int(time.time() * 1000)
        payload["token"] = payload.get("token", token)
        payload["rs"] = self._calc_rs()
        return payload

    def _post(self, endpoint: str, data: Dict[str, Any], k: str = "", token: str = "") -> Dict[str, Any]:
        payload = self._make_payload(data, k=k, token=token)
        headers = {
            "SIGN": self._calc_sign(payload),
            "X-Lang": self._lang,
            "Content-Type": "application/json;charset=UTF-8",
            "User-Agent": self.USER_AGENT,
            "Referer": f"{self.ORIGIN}/",
            "Origin": self.ORIGIN,
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9",
        }

        url = f"{self.BASE_API}{endpoint}"
        started = time.time()
        response = self._session.post(url, json=payload, headers=headers, timeout=30)
        duration_ms = int((time.time() - started) * 1000)
        response.raise_for_status()

        body = response.json()
        body_dict = body if isinstance(body, dict) else {}
        data_node = body_dict.get("data")
        list_count = 0
        data_total: Any = ""
        if isinstance(data_node, dict):
            data_total = data_node.get("total", "")
            if isinstance(data_node.get("lists"), list):
                list_count = len(data_node.get("lists", []))
        elif isinstance(data_node, list):
            list_count = len(data_node)
            data_total = len(data_node)

        self._request_logs.append(
            {
                "time": datetime.now().isoformat(timespec="seconds"),
                "endpoint": endpoint,
                "status": response.status_code,
                "api_code": body_dict.get("code", ""),
                "api_message": body_dict.get("message", ""),
                "duration_ms": duration_ms,
                "payload": payload,
                "response_data_total": data_total,
                "response_data_list_count": list_count,
            }
        )
        return body_dict

    def _bind_by_event_key(self, event_key: str) -> Tuple[Dict[str, Any], str]:
        candidates = [event_key]
        if event_key and not event_key.endswith("-views"):
            candidates.append(f"{event_key}-views")

        for idx, key in enumerate(candidates, 1):
            self._log(f"[{self._now()}] 绑定比赛尝试 {idx}/{len(candidates)}: {key}")
            detail = self._post("/match/detail", {"url_key": key, "is_encrypt": 2}, k="", token="")
            if detail.get("code") == 0:
                return detail, key
        raise RuntimeError("绑定失败：比赛链接无效，或比赛未开放观赛。")

    def _bind_by_k(self, k_value: str) -> Dict[str, Any]:
        self._log(f"[{self._now()}] URL 包含访问参数，尝试直接绑定。")
        detail = self._post("/match/detail", {"url_key": k_value, "is_encrypt": 1}, k=k_value, token="")
        if detail.get("code") == 0:
            return detail
        raise RuntimeError("绑定失败：访问参数无效或已失效。")

    def bind_match(self, user_input_url: str) -> MatchBindingInfo:
        self._request_logs.clear()
        normalized_url, event_key, k_query = self._parse_input_url(user_input_url)
        self._log(f"[{self._now()}] 开始绑定: {normalized_url}")

        bind_source = "event_key"
        if k_query:
            detail = self._bind_by_k(k_query)
            bind_source = "k_query"
        else:
            if not event_key:
                raise ValueError("URL 中没有比赛标识，也没有访问参数，无法绑定。")
            detail, used_key = self._bind_by_event_key(event_key)
            if used_key.endswith("-views") and used_key != event_key:
                self._log(f"[{self._now()}] 已自动回退到观赛地址: {used_key}")

        data = self._as_dict(detail.get("data"))
        k_value = str(data.get("url_key", "")).strip()
        if not k_value:
            raise RuntimeError("绑定失败：接口未返回有效访问参数。")

        token_resp = self._post("/common/account_token", {}, k=k_value, token="")
        token = str(self._as_dict(token_resp.get("data")).get("token", "")).strip()
        self._post("/match/surplus_time", {}, k=k_value, token=token)

        if not event_key:
            event_key = str(data.get("url_key", "")).strip()[:20] or "match"

        bind = MatchBindingInfo(
            input_url=user_input_url.strip(),
            normalized_url=normalized_url,
            event_key=event_key,
            k=k_value,
            token=token,
            title=str(data.get("title", "")),
            start_time=str(data.get("start_time", "")),
            end_time=str(data.get("end_time", "")),
            wp_stop_time=str(data.get("wp_stop_time", "")),
            login_url=f"https://match.ichunqiu.com/login?k={k_value}",
            situation_url=f"https://match.ichunqiu.com/situation?k={k_value}",
            problems_url=f"https://match.ichunqiu.com/situation/problems?k={k_value}",
            bind_source=bind_source,
            entry_type=self._to_int(data.get("entry_type"), 2),
            mode_type=self._to_int(data.get("mode_type"), 1),
            score_type=self._to_int(data.get("score_type"), 0),
            school_label=str(data.get("school_name") or self.OFFICIAL_CONTENT["school"]),
            industry_label=str(data.get("industry_name") or self.OFFICIAL_CONTENT["sshy"]),
        )

        self._log(f"[{self._now()}] 绑定成功: {bind.title}")
        self._log(f"[{self._now()}] 访问令牌 = {bind.k}")
        return bind

    def _payload_for_board(
        self,
        board_key: str,
        page_index: int,
        page_size: int,
        category_id: str = "",
    ) -> Dict[str, Any]:
        if board_key == "solved":
            return {
                "team_name": "",
                "industry_id": "",
                "attribute_id": "",
                "page_index": page_index,
                "page_size": page_size,
            }
        if board_key == "integral":
            return {
                "industry_id": "",
                "team_name": "",
                "page_index": page_index,
                "page_size": page_size,
            }
        if board_key == "question":
            return {
                "category_id": category_id,
                "page_index": page_index,
                "page_size": page_size,
            }
        if board_key == "single":
            return {
                "team_name": "",
                "page_index": page_index,
                "page_size": page_size,
            }
        if board_key == "category":
            return {
                "category_id": category_id,
                "team_name": "",
                "page_index": page_index,
                "page_size": page_size,
            }
        raise ValueError(f"不支持的榜单类型: {board_key}")

    def _row_unique_key(self, board_key: str, item: Dict[str, Any]) -> str:
        if board_key == "question":
            return str(
                item.get("question_id")
                or f"{item.get('title', '')}:{item.get('category_name', '')}:{item.get('type_id', '')}"
            )
        if board_key == "single":
            return str(
                item.get("user_id")
                or f"{item.get('user_name', '')}:{item.get('team_name', '')}:{item.get('user_rank', '')}"
            )
        return str(item.get("team_id") or item.get("source_id") or item.get("team_name") or "")

    def _collect_rows_for_board(
        self,
        bind: MatchBindingInfo,
        board_key: str,
        page_size: int,
        category_id: str = "",
    ) -> Tuple[str, List[Dict[str, Any]], int]:
        endpoint = self.BOARD_CONFIGS[board_key]["endpoint"]
        payload = self._payload_for_board(board_key, page_index=1, page_size=page_size, category_id=category_id)
        first_resp = self._post(endpoint, payload, k=bind.k, token=bind.token)
        if first_resp.get("code") != 0:
            message = str(first_resp.get("message", "未知错误"))
            board_name = self.BOARD_CONFIGS[board_key]["name"]
            # 业务级无数据（最常见的 code=116 "暂无数据"）→ 抛 NoDataError 让上层优雅处理
            if first_resp.get("code") == 116 or "暂无" in message or "没有" in message:
                raise NoDataError(board_name, message)
            raise RuntimeError(f"{board_name}接口调用失败: code={first_resp.get('code')} {message}")

        first_data = self._as_dict(first_resp.get("data"))
        first_rows = self._as_list(first_data.get("lists"))
        total = self._to_int(first_data.get("total"), len(first_rows))
        total_pages = max(1, math.ceil(total / page_size)) if total > 0 else 1

        rows: List[Dict[str, Any]] = [x for x in first_rows if isinstance(x, dict)]

        suffix = f" (category_id={category_id})" if board_key == "category" else ""
        self._log(f"[{self._now()}] 榜单接口: {endpoint}{suffix}")
        self._log(f"[{self._now()}] 预计总页数: {total_pages}")

        for page in range(2, total_pages + 1):
            page_payload = self._payload_for_board(
                board_key,
                page_index=page,
                page_size=page_size,
                category_id=category_id,
            )
            resp = self._post(endpoint, page_payload, k=bind.k, token=bind.token)
            if resp.get("code") != 0:
                self._log(f"[{self._now()}] 警告: 第 {page} 页返回异常 code={resp.get('code')}")
                continue
            data = self._as_dict(resp.get("data"))
            page_rows = self._as_list(data.get("lists"))
            rows.extend(x for x in page_rows if isinstance(x, dict))
            self._log(f"[{self._now()}] 已获取第 {page}/{total_pages} 页")

        dedup: List[Dict[str, Any]] = []
        seen: set[str] = set()
        for item in rows:
            key = self._row_unique_key(board_key, item)
            if key and key in seen:
                continue
            if key:
                seen.add(key)
            dedup.append(item)
        return endpoint, dedup, total_pages

    def _fetch_category_list(self, bind: MatchBindingInfo) -> List[Dict[str, Any]]:
        resp = self._post("/match/question/category", {}, k=bind.k, token=bind.token)
        if resp.get("code") != 0:
            message = resp.get("message", "未知错误")
            raise RuntimeError(f"获取题型列表失败: code={resp.get('code')} {message}")
        data = resp.get("data")
        if not isinstance(data, list):
            return []

        categories: List[Dict[str, Any]] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            category_id = item.get("id")
            title = str(item.get("title", "")).strip()
            if category_id in (None, ""):
                continue
            categories.append(
                {
                    "id": str(category_id),
                    "title": title or f"category_{category_id}",
                }
            )
        return categories

    @staticmethod
    def _flatten_row(item: Dict[str, Any], board_key: str) -> Dict[str, Any]:
        output: Dict[str, Any] = {}
        for key, value in item.items():
            if isinstance(value, (list, dict)):
                continue
            output[key] = value

        if board_key == "solved":
            solved_list = item.get("solved_list")
            if isinstance(solved_list, list):
                output["solved_count"] = len(solved_list)
                output["solved_titles"] = " | ".join(
                    str(x.get("title", "")).strip() for x in solved_list if isinstance(x, dict)
                )
        return output

    def _sort_dataframe(self, df: pd.DataFrame, board_key: str) -> pd.DataFrame:
        if df.empty:
            return df

        if board_key == "single" and "user_rank" in df.columns:
            work = df.copy()
            work["user_rank"] = pd.to_numeric(work["user_rank"], errors="coerce")
            sort_cols = [c for c in ["user_rank", "team_name", "user_name"] if c in work.columns]
            return work.sort_values(by=sort_cols, na_position="last") if sort_cols else work

        if "team_rank" in df.columns:
            work = df.copy()
            work["team_rank"] = pd.to_numeric(work["team_rank"], errors="coerce")
            sort_cols = [c for c in ["team_rank", "team_name"] if c in work.columns]
            return work.sort_values(by=sort_cols, na_position="last") if sort_cols else work

        return df

    def _translate_column_name(self, column_name: str, bind: MatchBindingInfo, board_key: str) -> str:
        if column_name == "team_name":
            key = "xsmc" if bind.entry_type == 1 else "teamName"
            return self.OFFICIAL_CONTENT.get(key, column_name)
        if column_name == "strong_point":
            if board_key == "single" or bind.entry_type == 1:
                return self.OFFICIAL_CONTENT.get("grqx", column_name)
            return self.OFFICIAL_CONTENT.get("dwqx", column_name)
        if column_name == "industry_team_rank":
            return f"{bind.industry_label}{self.OFFICIAL_CONTENT.get('pm', '排名')}"

        fixed_map = {
            "team_rank": "pm",
            "user_rank": "pm",
            "school": "school",
            "industry_name": "sshy",
            "total_score": "totalScore",
            "score_diff": "difference",
            "industry_score_diff": "difference",
            "title": "topicName",
            "category_name": "topicType",
            "theory_score": "aqzszf",
            "total_number": "jtsl",
            "solved_count": "jtsl",
            "solved_titles": "jcdtm",
            "user_name": "dwcy",
            "province_name": "province",
            "zone_name": "ssfq",
        }
        key = fixed_map.get(column_name)
        if key:
            return self.OFFICIAL_CONTENT.get(key, column_name)
        return column_name

    def _translate_dataframe_headers(self, df: pd.DataFrame, bind: MatchBindingInfo, board_key: str) -> pd.DataFrame:
        if df.empty:
            return df

        used: Dict[str, int] = {}
        new_columns: List[str] = []
        for column in df.columns:
            translated = self._translate_column_name(str(column), bind=bind, board_key=board_key)
            count = used.get(translated, 0) + 1
            used[translated] = count
            if count > 1:
                translated = f"{translated}_{count}"
            new_columns.append(translated)
        out = df.copy()
        out.columns = new_columns
        return out

    @staticmethod
    def _cell_display_len(value: Any) -> int:
        if value is None:
            return 0
        text = str(value)
        size = 0
        for ch in text:
            size += 2 if unicodedata.east_asian_width(ch) in {"W", "F"} else 1
        return size

    def _format_rank_sheet(self, ws: Worksheet) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(horizontal="center", vertical="center")
        header_font = Font(bold=True)

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row <= 0 or max_col <= 0:
            return

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = align
                if cell.row == 1:
                    cell.font = header_font

        for col in ws.iter_cols(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            letter = col[0].column_letter
            header = str(col[0].value or "").strip()
            max_len = max(self._cell_display_len(c.value) for c in col)
            if header in {"solved_titles", "解出的题目"}:
                width = min(max(72, max_len + 10), 220)
            else:
                width = min(max(8, max_len + 2), 90)
            ws.column_dimensions[letter].width = width

    def _fetch_team_roster(self, bind: MatchBindingInfo) -> List[Dict[str, Any]]:
        """拉取所有团队成员信息。先从积分榜获取全部 team_id，再逐队调用 /match/team/users。"""
        # 1) 拉所有 team_id（优先使用积分榜数据；若失败回退到解题榜）
        team_rows: List[Dict[str, Any]] = []
        for source_key in ("integral", "solved"):
            try:
                _, team_rows, _ = self._collect_rows_for_board(
                    bind=bind, board_key=source_key, page_size=200
                )
                if team_rows:
                    break
            except NoDataError:
                continue
            except Exception as exc:  # noqa: BLE001
                self._log(f"[{self._now()}] 获取团队列表({source_key})失败: {exc}")
                continue

        if not team_rows:
            raise NoDataError(self.BOARD_CONFIGS["team_info"]["name"], "未拉取到任何团队")

        self._log(f"[{self._now()}] 共 {len(team_rows)} 支队伍，开始逐队拉取成员")

        # 2) 逐队拉成员
        all_members: List[Dict[str, Any]] = []
        failed_teams: List[Tuple[str, str]] = []
        for team in team_rows:
            team_id = str(team.get("team_id") or team.get("source_id") or "").strip()
            if not team_id:
                continue
            payload = {"team_id": team_id, "page_index": 1, "page_size": 200}
            try:
                resp = self._post("/match/team/users", payload, k=bind.k, token=bind.token)
            except Exception as exc:  # noqa: BLE001
                failed_teams.append((team.get("team_name", team_id), str(exc)))
                continue
            if resp.get("code") != 0:
                msg = resp.get("message", "")
                if resp.get("code") == 116 or "暂无" in str(msg):
                    # 该队没有成员数据（单人/空队）—— 跳过即可
                    self._log(f"[{self._now()}] 队伍 {team.get('team_name', team_id)} 无成员数据")
                    continue
                failed_teams.append((team.get("team_name", team_id), f"code={resp.get('code')} {msg}"))
                continue
            data_node = self._as_dict(resp.get("data"))
            members = self._as_list(data_node.get("lists"))
            if not members:
                continue
            for m in members:
                if not isinstance(m, dict):
                    continue
                # 注入队伍维度信息：紧跟在 team_id 之后，便于阅读时按"队伍"聚类查看
                if "team_id" not in m or not m["team_id"]:
                    m["team_id"] = team_id
                team_dim = {
                    "team_name": team.get("team_name", ""),
                    "team_rank": team.get("team_rank", ""),
                    "school": team.get("school", ""),
                }
                ordered: Dict[str, Any] = {}
                inserted = False
                for k, v in m.items():
                    ordered[k] = v
                    if k == "team_id":
                        ordered.update(team_dim)
                        inserted = True
                if not inserted:
                    ordered.update(team_dim)
                all_members.append(ordered)

        if failed_teams:
            sample = ", ".join(f"{name}({err})" for name, err in failed_teams[:3])
            self._log(f"[{self._now()}] 部分队伍成员获取失败: {sample}")

        if not all_members:
            raise NoDataError(self.BOARD_CONFIGS["team_info"]["name"], "所有队伍均无成员数据")
        return all_members

    def _flatten_member_row(self, item: Dict[str, Any]) -> Dict[str, Any]:
        """成员行扁平化：剔除嵌套结构，把 user_type 转中文。"""
        output: Dict[str, Any] = {}
        for key, value in item.items():
            if isinstance(value, (list, dict)):
                continue
            output[key] = value
        if "user_type" in output:
            output["user_type"] = self._role_label(output["user_type"])
        return output

    def _build_dataframe(self, rows: List[Dict[str, Any]], bind: MatchBindingInfo, board_key: str) -> pd.DataFrame:
        flattened = [self._flatten_row(item, board_key=board_key) for item in rows]
        df = pd.DataFrame(flattened)
        if board_key == "solved" and "total_number" in df.columns and "solved_count" in df.columns:
            df = df.drop(columns=["solved_count"])
        df = self._sort_dataframe(df, board_key=board_key)
        df = self._translate_dataframe_headers(df, bind=bind, board_key=board_key)
        return df

    def export_board_xlsx(
        self,
        bind: MatchBindingInfo,
        output_root: str | Path,
        board_key: str,
        page_size: int = 50,
        debug_mode: bool = False,
    ) -> LeaderboardExportResult:
        if board_key not in self.BOARD_CONFIGS:
            raise ValueError(f"不支持的榜单类型: {board_key}")
        if page_size <= 0:
            page_size = 50

        board_name = self.BOARD_CONFIGS[board_key]["name"]
        self._request_logs.clear()
        self._log(f"[{self._now()}] 开始获取{board_name}...")

        out_root = Path(output_root).expanduser().resolve()
        out_root.mkdir(parents=True, exist_ok=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{self._safe_name(bind.event_key)}_{board_name}_{stamp}.xlsx"
        excel_path = out_root / file_name

        status = ExportStatus.SUCCESS
        error_message = ""
        total_rows = 0
        total_pages = 0
        endpoint = self.BOARD_CONFIGS[board_key]["endpoint"]
        used_sheet_names: set[str] = set()
        sheets_written = 0
        writer: Optional[pd.ExcelWriter] = None
        final_excel_path = ""

        try:
            writer = pd.ExcelWriter(excel_path, engine="openpyxl")
            try:
                if board_key == "team_info":
                    members = self._fetch_team_roster(bind)
                    total_rows = len(members)
                    flattened = [self._flatten_member_row(m) for m in members]
                    df = pd.DataFrame(flattened)
                    if not df.empty:
                        if "team_rank" in df.columns:
                            df["team_rank"] = pd.to_numeric(df["team_rank"], errors="coerce")
                        sort_cols = [c for c in ["team_rank", "team_name", "user_name"] if c in df.columns]
                        if sort_cols:
                            df = df.sort_values(by=sort_cols, na_position="last")
                        df = self._translate_dataframe_headers(df, bind=bind, board_key=board_key)
                    sheet_name = self._safe_sheet_name(self.BOARD_CONFIGS[board_key]["sheet"], used_sheet_names)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._format_rank_sheet(writer.book[sheet_name])
                    sheets_written += 1
                elif board_key != "category":
                    endpoint, rows, total_pages = self._collect_rows_for_board(
                        bind=bind,
                        board_key=board_key,
                        page_size=page_size,
                    )
                    total_rows = len(rows)
                    if total_rows == 0:
                        status = ExportStatus.NO_DATA
                        error_message = "该榜单暂无数据"
                        self._log(f"[{self._now()}] {board_name}无数据，跳过 Excel 生成")
                    else:
                        df = self._build_dataframe(rows, bind=bind, board_key=board_key)
                        sheet_name = self._safe_sheet_name(self.BOARD_CONFIGS[board_key]["sheet"], used_sheet_names)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self._format_rank_sheet(writer.book[sheet_name])
                        sheets_written += 1
                else:
                    categories = self._fetch_category_list(bind)
                    if not categories:
                        status = ExportStatus.NO_DATA
                        error_message = "题型榜单未返回分类"
                        self._log(f"[{self._now()}] 题型榜单无分类，跳过 Excel 生成")
                    else:
                        self._log(
                            f"[{self._now()}] 发现题型分类: {', '.join(str(c['title']) for c in categories)}"
                        )
                        for category in categories:
                            category_id = str(category["id"])
                            category_title = str(category["title"])
                            _, rows, pages = self._collect_rows_for_board(
                                bind=bind,
                                board_key=board_key,
                                page_size=page_size,
                                category_id=category_id,
                            )
                            total_rows += len(rows)
                            total_pages += pages
                            if rows:
                                df = self._build_dataframe(rows, bind=bind, board_key=board_key)
                                sheet_name = self._safe_sheet_name(category_title, used_sheet_names)
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                                self._format_rank_sheet(writer.book[sheet_name])
                                sheets_written += 1
                                self._log(
                                    f"[{self._now()}] 题型 {category_title} 导出完成: {len(rows)} 条, {pages} 页"
                                )
                        if sheets_written == 0:
                            status = ExportStatus.NO_DATA
                            error_message = "所有题型分类下均无数据"
                            self._log(f"[{self._now()}] 题型榜单各分类均无数据，跳过 Excel 生成")
                # 走到这里说明没抛异常，writer 已有内容可保存
                writer.close()
                writer = None
                final_excel_path = str(excel_path)
            except NoDataError as nd:
                status = ExportStatus.NO_DATA
                error_message = str(nd)
                self._log(f"[{self._now()}] {board_name}无数据: {error_message}")
                # NO_DATA 不再生成任何 Sheet / 文件
                try:
                    if writer is not None:
                        writer.close()
                except Exception:
                    pass
                writer = None
        except Exception as exc:  # noqa: BLE001
            status = ExportStatus.FAILED
            error_message = f"{type(exc).__name__}: {exc}"
            self._log(f"[{self._now()}] {board_name}失败: {error_message}")
        finally:
            if writer is not None:
                try:
                    writer.close()
                except Exception:
                    pass
            # 写出过 Sheet（SUCCESS）但路径未记录 → 补上；NO_DATA / FAILED 不补
            if status == ExportStatus.SUCCESS and sheets_written > 0 and not final_excel_path:
                final_excel_path = str(excel_path)
            # NO_DATA / FAILED：若磁盘上残留了空 xlsx（pd.ExcelWriter 已创建），清理掉
            if status != ExportStatus.SUCCESS and excel_path.exists():
                try:
                    excel_path.unlink()
                except Exception:
                    pass

        debug_json_path = ""
        if debug_mode:
            debug_path = out_root / f"{self._safe_name(bind.event_key)}_{board_name}_request_debug_{stamp}.json"
            try:
                debug_path.write_text(
                    json.dumps(
                        {
                            "bind": {
                                "title": bind.title,
                                "event_key": bind.event_key,
                                "k": bind.k,
                                "token": bind.token,
                                "entry_type": bind.entry_type,
                                "mode_type": bind.mode_type,
                                "score_type": bind.score_type,
                            },
                            "board_key": board_key,
                            "board_name": board_name,
                            "status": status,
                            "requests": self._request_logs,
                        },
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                debug_json_path = str(debug_path)
            except Exception as exc:  # noqa: BLE001
                self._log(f"[{self._now()}] 写入 debug json 失败: {exc}")

        if status == ExportStatus.SUCCESS:
            self._log(f"[{self._now()}] 导出完成: {final_excel_path}")
            self._log(f"[{self._now()}] 总条数: {total_rows}")
        elif status == ExportStatus.NO_DATA:
            self._log(f"[{self._now()}] {board_name}无数据，未生成 Excel 文件")
        else:
            self._log(f"[{self._now()}] {board_name}失败，未生成 Excel 文件")

        return LeaderboardExportResult(
            board_key=board_key,
            board_name=board_name,
            status=status,
            excel_path=final_excel_path,
            total_rows=total_rows,
            total_pages=total_pages if total_pages > 0 else 1,
            endpoint=endpoint,
            error_message=error_message,
            debug_json_path=debug_json_path,
        )

    def export_leaderboard_xlsx(
        self,
        bind: MatchBindingInfo,
        output_root: str | Path,
        page_size: int = 50,
        debug_mode: bool = False,
    ) -> LeaderboardExportResult:
        return self.export_board_xlsx(
            bind=bind,
            output_root=output_root,
            board_key="solved",
            page_size=page_size,
            debug_mode=debug_mode,
        )
